"""Read any spreadsheet file and output all sheets as plain text.
Supports: .xlsx, .xls, .xlsm, .xlsb, .csv, .tsv, .ods

Improvements over read-xlsx.py:
  - Properly handles merged cells (fills every merged cell with top-left value)
  - Uses public .charts property instead of private ._charts (future-proof)
  - pandas output not truncated (max_colwidth=None, max_columns=None)
  - No arbitrary row limit (original PowerShell fallback was capped at 200)

Usage: python read-xlsx.py <file_path>
"""
import sys
import os

if len(sys.argv) < 2:
    print("Usage: python read-xlsx.py <file_path>")
    sys.exit(1)

path = sys.argv[1]
ext = os.path.splitext(path)[1].lower()


def resolve_merged_cells(ws):
    """Return dict of (row, col) -> value for all cells inside merged ranges.

    openpyxl only stores the value in the top-left cell of a merge; all other
    cells read as None.  This function fills them so nothing gets silently lost.
    """
    merged = {}
    for merge_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged[(row, col)] = top_left_value
    return merged


def extract_xlsx_openpyxl(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    output = []
    for name in wb.sheetnames:
        ws = wb[name]
        merged = resolve_merged_cells(ws)
        output.append(f"=== Sheet: {name} ===")

        for row in ws.iter_rows():
            vals = []
            for cell in row:
                # Prefer merged-cell value; fall back to the cell's own value
                v = merged.get((cell.row, cell.column), cell.value)
                vals.append(str(v) if v is not None else "")
            if any(v.strip() for v in vals):
                output.append("\t".join(vals))

        # Use the public .charts property (replaces private ._charts)
        try:
            for ci, chart in enumerate(ws.charts):
                title = getattr(chart, 'title', None) or f"Chart {ci + 1}"
                # title may be a rich-text object
                if hasattr(title, 'text'):
                    title = title.text
                output.append(f"\n[Chart {ci + 1}: {title}]")
                try:
                    for series in chart.series:
                        ref = "N/A"
                        if hasattr(series, 'val') and series.val and hasattr(series.val, 'numRef'):
                            ref = str(series.val.numRef.f)
                        series_name = getattr(series, 'title', f"Series {ci + 1}")
                        output.append(f"  Series: {series_name} -- data ref: {ref}")
                except Exception:
                    output.append("  (chart series data: see cell ranges above)")
        except Exception:
            pass  # Chart extraction is non-critical

        output.append("")
    return "\n".join(output)


def extract_xlsx_pandas(path):
    import pandas as pd
    # Prevent column/value truncation that hides financial data
    pd.set_option('display.max_colwidth', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)

    xls = pd.ExcelFile(path)
    output = []
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, header=None)
        output.append(f"=== Sheet: {sheet} ===")
        output.append(df.to_string(index=False, header=False, na_rep=""))
        output.append("")
    return "\n".join(output)


# --- CSV / TSV: read directly, no library needed ---
if ext in ('.csv', '.tsv'):
    import csv
    delimiter = '\t' if ext == '.tsv' else ','
    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        for row in csv.reader(f, delimiter=delimiter):
            print('\t'.join(row))
    sys.exit(0)

# --- XLSX / XLSM: prefer openpyxl for merged cell awareness ---
if ext in ('.xlsx', '.xlsm'):
    try:
        print(extract_xlsx_openpyxl(path))
        sys.exit(0)
    except ImportError:
        sys.stderr.write("[info] openpyxl not installed, trying pandas...\n")
    except Exception as e:
        sys.stderr.write(f"[info] openpyxl failed ({e}), trying pandas...\n")

# --- All remaining formats (.xls, .xlsb, .ods) and openpyxl fallback ---
try:
    print(extract_xlsx_pandas(path))
    sys.exit(0)
except ImportError:
    print(f"[ERROR] Cannot read {ext}: install openpyxl and/or pandas\n  pip install openpyxl pandas")
    sys.exit(1)
except Exception as e:
    print(f"[ERROR] Failed to read {path}: {e}")
    sys.exit(1)
